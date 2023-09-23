import csv
import pandas as pd
from openpyxl import load_workbook as lowk

# 向csv中写入指定数据
def writer_csv(file_path, datas):
    try:
        files = open(file_path, "a", encoding="utf-8", newline="")
        csvwr = csv.writer(files, dialect="excel")
        csvwr.writerow(datas)
        print("写入完成")
    except Exception as e:
        print('写入失败',e)
    

#向excel中追加数据
def write_excel(file_path,datas):
    try:
        wookbook = lowk(filename=file_path)
        sheet = wookbook.active
        sheet.append(datas)
        wookbook.save(filename=file_path)
        print('写入成功')
    except Exception as e:
        print('写入失败',e)
    

# for i in range(0,10):
#     datas = [fake.name(),fake.date(pattern="%Y-%m-%d", end_datetime=None),
#         fake.phone_number(), fake.address(), fake.city_name()]
#     write_excel('./datas/test.xlsx',datas=datas)
# wookbook = lowk(filename='./datas/test.xlsx')
# sheet = wookbook.active



# dd = []
# for datas in sheet.iter_rows(min_row=2, max_row=sheet.max_row):
#     d=[]
#     for i in datas:
#         d.append(i.value)
#         print(i.value)
#     dd.append(d)
    
# print(dd)